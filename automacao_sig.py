import pandas as pd
import os
import time
import datetime
import openpyxl
import logging
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

# Carrega variáveis
load_dotenv()

# --- CONFIGURAÇÕES ---
NOME_ARQUIVO_EXCEL = os.getenv("NOME_ARQUIVO_EXCEL") 
NOME_ARQUIVO_HISTORICO = os.getenv("NOME_ARQUIVO_HISTORICO") 
URL_DESPESA = "https://sig.loopbrasil.com/remocao/cadastro-despesa"
ABA_CALCULOS = "Calculos"

# --- FUNÇÕES AUXILIARES ---

def marcar_lancamento_excel(placa_alvo, nome_coluna_alvo):
    logging.info(f"    💾 Atualizando Excel para {placa_alvo} na coluna '{nome_coluna_alvo}'...")
    try:
        wb = openpyxl.load_workbook(NOME_ARQUIVO_EXCEL)
        if ABA_CALCULOS not in wb.sheetnames: return
        ws = wb[ABA_CALCULOS]
        
        header_row = None
        col_placa_idx = None
        col_alvo_idx = None
        
        for r in range(1, 6):
            vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
            if "Placa" in vals and nome_coluna_alvo in vals:
                header_row = r
                col_placa_idx = vals.index("Placa") + 1
                col_alvo_idx = vals.index(nome_coluna_alvo) + 1
                break
        
        if not col_placa_idx or not col_alvo_idx: return

        encontrou = False
        for row in ws.iter_rows(min_row=header_row + 1):
            cell_placa = row[col_placa_idx - 1]
            if str(cell_placa.value).strip().upper() == str(placa_alvo).strip().upper():
                row[col_alvo_idx - 1].value = "Sim"
                encontrou = True
                break
        
        if encontrou:
            wb.save(NOME_ARQUIVO_EXCEL)
            logging.info("    ✅ Excel Salvo com Sucesso.")
    except Exception as e:
        logging.error(f"    ERRO AO GRAVAR EXCEL: {e}")

def limpar_valor_para_float(valor):
    if pd.isna(valor) or str(valor).strip() == "": return 0.0
    try:
        if isinstance(valor, (int, float)): return float(valor)
        texto = str(valor).strip().replace("R$", "").replace(" ", "")
        if "," in texto: texto = texto.replace(".", "").replace(",", ".")
        return float(texto)
    except: return 0.0

def formatar_valor_para_site(valor_float):
    try: return f"{valor_float:.2f}".replace(".", ",")
    except: return "0,00"

def extrair_cidade_estado(patio_texto):
    try:
        if "-" in str(patio_texto):
            partes = str(patio_texto).split("-")
            estado = partes[-1].strip()
            cidade = "-".join(partes[:-1]).strip()
            return cidade, estado
        return str(patio_texto), "SP"
    except: return "", ""

def limpar_data_para_site(valor):
    hoje_str = datetime.date.today().strftime("%d/%m/%Y")
    
    if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nat":
        return hoje_str
    
    try:
        if isinstance(valor, (datetime.date, datetime.datetime, pd.Timestamp)):
            return valor.strftime("%d/%m/%Y")
        
        texto = str(valor).strip()
        texto = texto.split(' ')[0].split('T')[0]
        
        if "-" in texto:
            partes = texto.split("-")
            if len(partes[0]) == 4:
                dt = datetime.datetime.strptime(texto, "%Y-%m-%d")
                return dt.strftime("%d/%m/%Y")
            elif len(partes[2]) == 4:
                return texto.replace("-", "/")
        if "/" in texto:
            return texto
        return hoje_str
    except:
        return hoje_str

def selecionar_opcao_parcial(driver, xpath, texto_parcial):
    try:
        select_elem = driver.find_element(By.XPATH, xpath)
        select_obj = Select(select_elem)
        texto_buscado = str(texto_parcial).strip().upper()
        try:
            select_obj.select_by_visible_text(texto_buscado)
            return True
        except: pass
        for opt in select_obj.options:
            if texto_buscado in opt.text.upper():
                select_obj.select_by_visible_text(opt.text)
                return True
        logging.warning(f"    AVISO: Opção '{texto_buscado}' não encontrada (Parcial).")
        return False
    except: return False

def preencher_formulario(driver, dados, modo):
    """
    Preenche e SALVA automaticamente.
    """
    XPATHS = {
        "patio": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[1]/div[1]/div[1]/div/select",
        "comitente": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[1]/div[2]/div[1]/div/select",
        "dt_solic": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[1]/div[3]/input",
        "dt_real": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[2]/div[1]/input",
        "dt_final": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[2]/div[2]/input",
        "tipo_despesa": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[2]/div[3]/div[1]/div/select",
        "placa": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[2]/div[4]/input",
        "vinculo_placa": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[2]/div[5]/div[1]/div/select",
        "tipo_apreensao": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[5]/div[1]/div[1]/div/select",
        "guincheiro": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[5]/div[2]/div[1]/div/select",
        "valor": "/html/body/div[3]/div/div/div/form/div/div/fieldset[1]/div[5]/div[3]/input",
        "uf_origem": "/html/body/div[3]/div/div/div/form/div/div/fieldset[2]/fieldset[1]/div/div[1]/div[2]/div[1]/div/select",
        "cid_origem": "/html/body/div[3]/div/div/div/form/div/div/fieldset[2]/fieldset[1]/div/div[1]/div[3]/input",
        "uf_destino": "/html/body/div[3]/div/div/div/form/div/div/fieldset[2]/fieldset[2]/div/div[1]/div[2]/div[1]/div/select",
        "cid_destino": "/html/body/div[3]/div/div/div/form/div/div/fieldset[2]/fieldset[2]/div/div[1]/div[3]/input",
        "salvar": "/html/body/div[3]/div/div/div/form/div/div/div/div[1]/button"
    }

    logging.info(f"    > Preenchendo formulário: {modo}")
    
    # 1. Pátio
    patio_texto = dados['patio_nome'].strip()
    try:
        select_patio = Select(driver.find_element(By.XPATH, XPATHS["patio"]))
        try:
            select_patio.select_by_visible_text(patio_texto)
        except:
            encontrou_similar = False
            for opt in select_patio.options:
                if patio_texto.upper() in opt.text.upper():
                    select_patio.select_by_visible_text(opt.text)
                    encontrou_similar = True
                    break
            if not encontrou_similar:
                logging.error(f"    ERRO CRÍTICO PÁTIO: '{patio_texto}' não encontrado.")
                return False
    except: return False
    
    # 2. Comitente e Datas
    Select(driver.find_element(By.XPATH, XPATHS["comitente"])).select_by_visible_text("AYMORE CREDITO FINANCIAMENTO E INVESTIMENTO")
    
    driver.find_element(By.XPATH, XPATHS["dt_solic"]).clear()
    driver.find_element(By.XPATH, XPATHS["dt_solic"]).send_keys(dados['dt_solic'])
    driver.find_element(By.XPATH, XPATHS["dt_real"]).clear()
    driver.find_element(By.XPATH, XPATHS["dt_real"]).send_keys(dados['dt_real'])
    driver.find_element(By.XPATH, XPATHS["dt_final"]).clear()
    driver.find_element(By.XPATH, XPATHS["dt_final"]).send_keys(dados['dt_final'])

    # 3. Tipo, Placa, Vinculo
    tipo_txt = "Frete Restituição" if modo == "RECEITA" else "Restituição Judicial"
    Select(driver.find_element(By.XPATH, XPATHS["tipo_despesa"])).select_by_visible_text(tipo_txt)
    
    el_placa = driver.find_element(By.XPATH, XPATHS["placa"])
    el_placa.clear()
    el_placa.send_keys(dados['placa'])
    
    time.sleep(2.5) 
    try: Select(driver.find_element(By.XPATH, XPATHS["vinculo_placa"])).select_by_index(1)
    except:
        logging.error("    ERRO: Contrato não carregou.")
        return False

    # 4. Guincheiro e Valor
    selecionar_opcao_parcial(driver, XPATHS["guincheiro"], dados['transportadora'])
    valor_site = formatar_valor_para_site(dados['valor_float'])
    driver.find_element(By.XPATH, XPATHS["valor"]).send_keys(valor_site)

    # 5. Despesa Específicos
    if modo == "DESPESA":
        Select(driver.find_element(By.XPATH, XPATHS["tipo_apreensao"])).select_by_visible_text("Judicial")
        cidade_pt, uf_pt = extrair_cidade_estado(dados['patio_nome'])
        Select(driver.find_element(By.XPATH, XPATHS["uf_origem"])).select_by_visible_text(uf_pt)
        driver.find_element(By.XPATH, XPATHS["cid_origem"]).send_keys(cidade_pt + Keys.TAB)
        Select(driver.find_element(By.XPATH, XPATHS["uf_destino"])).select_by_visible_text(uf_pt)
        driver.find_element(By.XPATH, XPATHS["cid_destino"]).send_keys(dados['cidade_destino'] + Keys.TAB)

    logging.info(f"    Campos preenchidos. Valor: {valor_site}")
    
    # --- CLIQUE FINAL AUTOMÁTICO ---
    logging.info("    🚀 Salvando lançamento automaticamente...")
    driver.find_element(By.XPATH, XPATHS["salvar"]).click()
    
    time.sleep(4) # Pausa para processamento do servidor
    return True

# --- FUNÇÃO PRINCIPAL ---
def iniciar_automacao_sig():
    logging.info("--- Iniciando Módulo SIG (Automático/Headless) ---")
    
    try:
        df_status = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=ABA_CALCULOS)
        df_dados = pd.read_excel(NOME_ARQUIVO_HISTORICO)
        
        df_status['Placa'] = df_status['Placa'].astype(str).str.strip().str.upper()
        df_dados['Placa'] = df_dados['Placa'].astype(str).str.strip().str.upper()
        
        cols_para_ler = [
            'Placa', 'Pátio', 'Transportadora', 'Calculo_cobrança', 
            'Valor_Base_Guincho2', 'Cidade convertida', 'Teste',
            'Data_Restituicao', 'Fechamento_Solicitacao'
        ]
        cols_existentes = [c for c in cols_para_ler if c in df_dados.columns]
        df_merged = pd.merge(df_status, df_dados[cols_existentes], on='Placa', how='left')
        
        logging.info(f"Base SIG carregada: {len(df_merged)} linhas.")
    except Exception as e:
        logging.error(f"Erro ao ler planilhas SIG: {e}")
        return

    # CONFIGURAÇÃO HEADLESS (INVISÍVEL)
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--no-sandbox")
    
    driver = None
    try: 
        driver = webdriver.Chrome(options=chrome_options)
    except: 
        logging.error("Erro ao iniciar Driver SIG. Verifique o chromedriver.")
        return

    try:
        logging.info("Acessando SIG e realizando Login...")
        driver.get("https://sig.loopbrasil.com")
        usuario = os.getenv('SIG_LOGIN')
        senha = os.getenv('SIG_PASSWORD')
        time.sleep(1)
        driver.find_element(By.XPATH, "/html/body/div[2]/form/input[2]").send_keys(usuario)
        driver.find_element(By.XPATH, "/html/body/div[2]/form/input[3]").send_keys(senha)
        driver.find_element(By.XPATH, "/html/body/div[2]/form/input[3]").send_keys(Keys.RETURN)
        time.sleep(4)
        logging.info("Login realizado.")

        for idx, row in df_merged.iterrows():
            placa = row['Placa']
            status_atual = str(row['Status atual']).strip()
            
            if status_atual.upper() != "RESTITUIÇÃO CONCLUÍDA": continue 
            
            logging.info(f"\n🚙 SIG Analisando: {placa}")
            
            try: flag_teste = int(float(row.get('Teste', 1)))
            except: flag_teste = 1
            
            transp_original = str(row['Transportadora'])
            dt_rest_limpa = limpar_data_para_site(row.get('Data_Restituicao'))
            dt_fech_limpa = limpar_data_para_site(row.get('Fechamento_Solicitacao'))

            dados_placa = {
                'placa': placa,
                'patio_nome': str(row['Pátio']),
                'transportadora': transp_original, 
                'cidade_destino': str(row['Cidade convertida'])
            }
            
            # --- RECEITA ---
            if str(row['Lançado receita?']).strip() == "Não":
                val = limpar_valor_para_float(row['Calculo_cobrança'])
                if val > 0:
                    dados_placa['valor_float'] = val
                    if flag_teste == 0: dados_placa['transportadora'] = "LOOP - LOOP BRASIL"
                    else: dados_placa['transportadora'] = transp_original
                    
                    dados_placa['dt_solic'] = dt_fech_limpa
                    dados_placa['dt_real']  = dt_fech_limpa
                    dados_placa['dt_final'] = dt_fech_limpa

                    driver.get(URL_DESPESA)
                    time.sleep(1.5)
                    
                    if preencher_formulario(driver, dados_placa, modo="RECEITA"):
                        marcar_lancamento_excel(placa, "Lançado receita?")

            # --- DESPESA ---
            if str(row['Lançado despesa?']).strip() == "Não":
                val = limpar_valor_para_float(row['Valor_Base_Guincho2'])
                if val > 0:
                    dados_placa['valor_float'] = val
                    dados_placa['transportadora'] = transp_original 
                    
                    dados_placa['dt_solic'] = dt_rest_limpa
                    dados_placa['dt_real']  = dt_fech_limpa
                    dados_placa['dt_final'] = dt_fech_limpa
                    
                    driver.get(URL_DESPESA)
                    time.sleep(1.5)
                    
                    if preencher_formulario(driver, dados_placa, modo="DESPESA"):
                        marcar_lancamento_excel(placa, "Lançado despesa?")

        logging.info("\n🏁 Módulo SIG finalizado.")

    except Exception as e:
        logging.error(f"Erro Crítico SIG: {e}")
    finally:
        if driver: driver.quit()

if __name__ == "__main__":
    # Apenas para teste isolado
    configurar_logger_dinamico = lambda: None 
    iniciar_automacao_sig()