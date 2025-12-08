import pandas as pd
import os
import time
import json
import base64
import datetime
import logging
import telegram
import asyncio
import glob
import math
import unicodedata
import re
import shutil
import win32com.client as win32
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

load_dotenv()

# --- Configurações ---
NOME_ARQUIVO_EXCEL = os.getenv("NOME_ARQUIVO_EXCEL","Base_Restituicoes.xlsx")
NOME_ABA_CALCULOS = "Calculos"
NOME_ABA_BASES = "Bases"
PASTA_DOWNLOADS = os.getenv("PASTA_DOWNLOADS")
NOME_ARQUIVO_HISTORICO = os.getenv("NOME_ARQUIVO_HISTORICO","historico_processamento.xlsx")

CAMINHO_BASE_EXTERNA = os.getenv("CAMINHO_BASE_EXTERNA", "remocao-restituicao.xlsx")
CAMINHO_CUSTO_RESTITUICAO = os.getenv("CAMINHO_CUSTO_RESTITUICAO", "Custo_Restituicao.xlsx")

COLUNAS_EXTERNAS_MAP = {
    'Placa': 'placa_key',
    'Guincheiro': 'transp_raw',
    'nm': 'patio_raw',
    'CidadeOrigem': 'cidade_raw',
    'financiado': 'financiado_db',
    'cpf': 'cpf_db',
    'Contrato': 'contrato_externo',
    'ValorGuincheiro': 'valor_base_db',
    'DataSolicitacao': 'Data de Remoção',
    'Marca': 'Marca',
    'Modelo': 'Modelo',
    'Categoria': 'Categoria_Ext',
    'Chassi': 'Chassi'
}

COLUNA_PLACA = "Placa"
COLUNA_CONTRATO = "Contrato"
COLUNA_CATEGORIA = "Categoria"
COLUNA_TESTE = "Teste"
COLUNA_END1 = "Endereço transportadora"
COLUNA_END2 = "Endereço Pátio"
COLUNA_END3 = "Cidade convertida"
COLUNA_STATUS_SAFEDOC = "Conferencia SafeDoc"

URL_BANCO = os.getenv("URL_BANCO")
USUARIO_BANCO = os.getenv("USUARIO_BANCO")
SENHA_BANCO = os.getenv("SENHA_BANCO")

VALOR_RANGES = {
    "leve": [(200, 241), (500, 468), (700, 620), (1000, 900), (9999, 1320)],
    "moto": [(200, 230), (500, 438), (700, 580), (1000, 795), (9999, 880)],
    "pesado": [(200, 665), (500, 1045), (700, 2020), (1000, 3235), (9999, 4175)]
}

# --- SELETORES ---
SELECTORS = {
    "google_maps": {
        "km_xpaths_list": [
            "/html/body/div[1]/div[2]/div[9]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[5]/div[1]/div[1]/div/div[1]/div[2]/div",
            "/html/body/div[1]/div[3]/div[9]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[5]/div[1]/div[1]/div/div[1]/div[2]/div",
            "//div[contains(@id, 'section-directions-trip-0')]//div[contains(text(), 'km')]",
            "//div[contains(@class, 'ivN21e')]"
        ],
        "canvas_map": "canvas"
    },
    "login": {
        "usuario": "/html/body/div/main/div/div[2]/form/div/div[2]/div/input",
        "senha": "/html/body/div/main/div/div[2]/form/div/div[3]/div/input", 
        "botao": "/html/body/div/main/div/div[2]/form/div/div[4]/input[1]"
    },
    "gca_menu": {
        "link_1": "/html/body/main/section/form/div/div/div/div/div/div/div[1]/div[2]/div/nobr/div/div[1]/span[1]/img",
        "link_2": "/html/body/main/section/form/div/div/div/div/div/div/div[1]/div[2]/div/nobr/div/div[2]/div[3]/span[1]/img[2]",
        "link_3": "/html/body/main/section/form/div/div/div/div/div/div/div[1]/div[2]/div/nobr/div/div[2]/div[4]/div[2]/img"
    },
    "iframes": {
        "externo": "ifrmForm",
        "interno": "ifrmObject"
    },
    "form_upload": {
        "input_arquivo": "_ctl0_ContentPlaceHolder_IDX_FILE",
        "select_status": "/html/body/form/div/div/div/div/div[2]/div[3]/select",
        "input_data": "/html/body/form/div/div/div/div/div[2]/div[4]/div/input",
        "input_contrato": "/html/body/form/div/div/div/div/div[2]/div[5]/input",
        "input_placa": "/html/body/form/div/div/div/div/div[2]/div[6]/input",
        "select_tipo_despesa": "/html/body/form/div/div/div/div/div[2]/div[7]/select",
        "input_valor": "/html/body/form/div/div/div/div/div[2]/div[8]/input",
        "input_caixa_arquivo": "/html/body/form/div/div/div/div/div[2]/div[9]/input",
        "input_observacao": "/html/body/form/div/div/div/div/div[2]/div[10]/input",
        "botao_salvar": "/html/body/form/div/div/div/div/div[3]/input",
        "mensagem_sucesso": "/html/body/form/div/div/div/div/div[4]/div/span"
    }
}

# --- Funções de Suporte ---
def configurar_logger_dinamico():
    try: diretorio_script = os.path.dirname(os.path.abspath(__file__))
    except: diretorio_script = os.getcwd()
    
    pasta_logs_raiz = os.path.join(diretorio_script, ".logs")
    hoje_str = datetime.date.today().strftime("%Y-%m-%d")
    pasta_diaria = os.path.join(pasta_logs_raiz, hoje_str)
    os.makedirs(pasta_diaria, exist_ok=True)
    
    padrao = os.path.join(pasta_diaria, f"log_{hoje_str}_v*.txt")
    maior = 0
    for arq in glob.glob(padrao):
        try: maior = max(maior, int(os.path.splitext(os.path.basename(arq))[0].split("_v")[-1]))
        except: pass
    nome_log = os.path.join(pasta_diaria, f"log_{hoje_str}_v{maior + 1}.txt")

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    if logger.hasHandlers():
        logger.handlers.clear()

    file_handler = logging.FileHandler(nome_log, mode='w', encoding='utf-8')
    formatter_file = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    file_handler.setFormatter(formatter_file)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    formatter_console = logging.Formatter('%(message)s') 
    console_handler.setFormatter(formatter_console)
    logger.addHandler(console_handler)

    print(f"\n📂 Arquivo de Log criado: {os.path.basename(pasta_diaria)}/{os.path.basename(nome_log)}\n")

def limpar_texto_estilo_excel(texto):
    if not isinstance(texto, str): return ""
    nfkd = unicodedata.normalize('NFKD', texto)
    texto_sem_acento = "".join([c for c in nfkd if not unicodedata.combining(c)])
    return " ".join(re.sub(r'[^A-Z0-9\s]', '', texto_sem_acento.upper()).split())

def formatar_moeda_br(valor):
    try:
        if valor is None or str(valor).strip() == "": return "0,00"
        if isinstance(valor, str) and ',' in valor and '.' not in valor: return valor
        
        val_str = str(valor).replace(',', '.')
        val_float = float(val_str)
        
        return f"{val_float:.2f}".replace('.', ',')
    except:
        return "0,00"

def formatar_data_ptbr(valor):
    if pd.isna(valor) or str(valor).strip() in ['', 'nan', 'None', 'NaT']: return ""
    
    if hasattr(valor, 'strftime'): 
        return valor.strftime('%d/%m/%Y')
    
    try:
        if isinstance(valor, str) and '-' in valor:
            dt = pd.to_datetime(valor, errors='coerce')
        else:
            dt = pd.to_datetime(valor, errors='coerce', dayfirst=True)
            
        if pd.notna(dt): 
            return dt.strftime('%d/%m/%Y')
    except: pass
    
    return str(valor)

def carregar_bases_de_enderecos():
    try:
        df_bases = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=NOME_ABA_BASES, header=None)
        dict_transp = dict(zip(df_bases[0].dropna().astype(str).str.strip().str.upper(), df_bases[1].dropna().astype(str).str.strip()))
        dict_patio = dict(zip(df_bases[2].dropna().astype(str).apply(limpar_texto_estilo_excel), df_bases[3].dropna().astype(str).str.strip()))
        return dict_transp, dict_patio
    except Exception as e:
        logging.error(f"Erro Base Endereços: {e}")
        return {}, {}

def carregar_base_externa_rede():
    logging.info("Carregando base...")
    nome_arquivo_rede = os.path.basename(CAMINHO_BASE_EXTERNA)
    caminho_local_direto = os.path.join(os.getcwd(), nome_arquivo_rede)
    caminho_final = CAMINHO_BASE_EXTERNA
    usando_copia = False
    
    if os.path.exists(caminho_local_direto):
        logging.info(f"USANDO ARQUIVO LOCAL: {caminho_local_direto}")
        caminho_final = caminho_local_direto
    elif os.path.exists(CAMINHO_BASE_EXTERNA):
        if "\\" in CAMINHO_BASE_EXTERNA or "//" in CAMINHO_BASE_EXTERNA:
            try:
                logging.info("Copiando arquivo da rede para local temporário...")
                caminho_temp = os.path.join(os.getcwd(), "temp_rede_copy.xlsx")
                shutil.copy2(CAMINHO_BASE_EXTERNA, caminho_temp)
                caminho_final = caminho_temp
                usando_copia = True
            except Exception as e:
                logging.error(f"Erro cópia rede: {e}")
    else:
        logging.critical(f"Arquivo não encontrado: {CAMINHO_BASE_EXTERNA}")
        return pd.DataFrame()

    try:
        df_ext = pd.read_excel(caminho_final, sheet_name="remocao", usecols=list(COLUNAS_EXTERNAS_MAP.keys()), engine='openpyxl', dtype=str)
        df_ext.rename(columns=COLUNAS_EXTERNAS_MAP, inplace=True)
        if 'placa_key' in df_ext.columns:
            df_ext['placa_key'] = df_ext['placa_key'].str.strip().str.upper()
            df_ext.drop_duplicates(subset=['placa_key'], inplace=True)
        
        logging.info(f"Base Externa carregada com {len(df_ext)} linhas.")
        if usando_copia and os.path.exists(caminho_final): os.remove(caminho_final)
        return df_ext
    except Exception as e:
        logging.critical(f"Erro leitura externa: {e}")
        return pd.DataFrame()

def carregar_tabela_custos_jpr():
    if not os.path.exists(CAMINHO_CUSTO_RESTITUICAO): return {}
    try:
        df = pd.read_excel(CAMINHO_CUSTO_RESTITUICAO, sheet_name='Todos', header=None, skiprows=1)
        tabela_jpr = {}
        for _, row in df.iterrows():
            cid = limpar_texto_estilo_excel(str(row[1]))
            pat = limpar_texto_estilo_excel(str(row[2]))
            tra = limpar_texto_estilo_excel(str(row[3]))
            tabela_jpr[(cid, pat, tra)] = {'Moto': row[4], 'Leve': row[5], 'Caminhonete': row[6]}
        return tabela_jpr
    except: return {}

def calcular_cobranca_individual(tipo_lib, tipo_restituicao, v_base, v_rem, v_base2):

    try:
        def to_float(val):
            if isinstance(val, (int, float)): return float(val)
            try:
                clean = str(val).replace(',', '.').strip()
                return float(clean) if clean else 0.0
            except:
                return 0.0

        v_base = to_float(v_base)
        v_rem = to_float(v_rem)
        v_base2 = to_float(v_base2)
        tipo_lib = str(tipo_lib).strip()
        
        teste = 1 if str(tipo_restituicao).strip() == "Transportadora" else 0

        resultado = 0.0

        if tipo_lib == "Determinação Judicial" and teste == 1:
            resultado = 0.0
        
        elif tipo_lib == "Acordo" and teste == 1:
            resultado = ((v_base - v_rem) + v_base2) * 1.15
        
        elif tipo_lib == "Acordo" and teste == 0:
            resultado = (v_base - v_rem) * 1.15
        
        return max(0.0, resultado)

    except Exception as e:
        logging.error(f"Erro no cálculo individual: {e}")
        return 0.0

def sincronizar_dados_dinamicos_local(df_historico, df_ext):
    logging.info("Sincronizando: Local -> Rede -> Histórico...")
    try:
        try:
            cols_local = [
                COLUNA_PLACA, 
                'Status atual', 
                'Fechamento Solicitação', 
                'Tipo de liberação', 
                'Tipo de restituição', 
                'Data Restituição', 
                'Conferencia SafeDoc' 
            ]
            
            df_local_raw = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=NOME_ABA_CALCULOS, dtype=str)
            cols_ex = [c for c in cols_local if c in df_local_raw.columns]
            df_local = df_local_raw[cols_ex].copy()
            for col in df_local.columns: df_local[col] = df_local[col].astype(str).str.strip()
            df_local.drop_duplicates(subset=[COLUNA_PLACA], inplace=True)
            placas_locais = df_local[COLUNA_PLACA].unique()
            df_local_idx = df_local.set_index(COLUNA_PLACA).to_dict('index')
        except: return df_historico

        df_ext_idx = df_ext.set_index('placa_key').to_dict('index') if not df_ext.empty else {}
        
        mapa_estatico = {
            'cpf_db': 'CPF_Banco', 'financiado_db': 'Financiado_Banco', 
            'contrato_externo': 'Contrato_Externo', 'valor_base_db': 'Valor_Base_Guincho', 
            'transp_raw': 'Transportadora', 'patio_raw': 'Pátio', 
            'cidade_raw': 'Cidade convertida', 'Data de Remoção': 'Data de Remoção', 
            'Marca': 'Marca', 'Modelo': 'Modelo', 
            'Categoria_Ext': 'Categoria', 'Chassi': 'Chassi'
        }
        
        mapa_dinamico = {
            'Status atual': 'Status_Atual', 
            'Fechamento Solicitação': 'Fechamento_Solicitacao', 
            'Tipo de liberação': 'Tipo_Liberacao', 
            'Tipo de restituição': 'Tipo_Restituicao', 
            'Data Restituição': 'Data_Restituicao',
            'Conferencia SafeDoc': 'Conferencia SafeDoc' 
        }

        for c in list(mapa_estatico.values()) + list(mapa_dinamico.values()):
            if c not in df_historico.columns: df_historico[c] = None

        novas = []
        alt_count = 0
        for placa in placas_locais:
            if not placa or str(placa).lower() in ['nan', 'none', '']: continue
            rede = df_ext_idx.get(placa, {})
            local = df_local_idx.get(placa, {})
            mask = df_historico[COLUNA_PLACA] == placa
            
            if mask.any():
                idx = df_historico[mask].index[0]
                
                for cr, ch in mapa_estatico.items():
                    val_h = str(df_historico.at[idx, ch]).strip()
                    val_n = str(rede.get(cr, '')).strip()
                    if val_h in ['nan', 'None', '', 'NaT'] and val_n not in ['nan', 'None', '']: 
                        df_historico.at[idx, ch] = val_n
                
                for cl, ch in mapa_dinamico.items():
                    if cl not in local: continue
                    vh = str(df_historico.at[idx, ch]).strip()
                    vl = str(local.get(cl, '')).strip()
                    
                    if vl not in ['nan', 'None', '']:
                        if vh in ['nan', 'None', '', 'NaT']:
                            df_historico.at[idx, ch] = vl
                            alt_count += 1
                        elif vh != vl:
                            if ch == 'Conferencia SafeDoc':
                                df_historico.at[idx, ch] = vl
                                alt_count += 1
                            else:
                                print(f"\n>>> CONFLITO {placa} ({ch}): Hist='{vh}' vs Local='{vl}'")
                                resp = input("    Atualizar? (S/N): ").strip().upper()
                                if resp == 'S': 
                                    df_historico.at[idx, ch] = vl
                                    alt_count += 1
                                    print("    [ATUALIZADO]")
            else:
                novo = {COLUNA_PLACA: placa}
                for cr, ch in mapa_estatico.items(): novo[ch] = rede.get(cr, None)
                for cl, ch in mapa_dinamico.items(): novo[ch] = local.get(cl, None)
                novas.append(novo)

        if novas: df_historico = pd.concat([df_historico, pd.DataFrame(novas)], ignore_index=True)
        if 'Tipo_Restituicao' in df_historico.columns:
            df_historico['Teste'] = df_historico['Tipo_Restituicao'].astype(str).str.strip().apply(
                lambda x: 1 if x == "Transportadora" else 0
            )
        else:
            df_historico['Teste'] = 0
        logging.info(f"Sincronização: {alt_count} alterações, {len(novas)} novos.")
        return df_historico
    except Exception as e: 
        logging.error(f"Erro na sincronização: {e}")
        return df_historico

def calcular_valor_restituicao_final(transp_nome, cidade_nome, patio_nome, categoria, valor_remocao, tabela_jpr):
    if 'JPR' not in transp_nome.upper(): return valor_remocao
    chave = (limpar_texto_estilo_excel(cidade_nome), limpar_texto_estilo_excel(patio_nome), limpar_texto_estilo_excel(transp_nome))
    valores = tabela_jpr.get(chave)
    if not valores: return "Não encontrada"
    cat_lower = str(categoria).strip().lower()
    if 'moto' in cat_lower: return valores.get('Moto', 0)
    elif 'leve' in cat_lower: return valores.get('Leve', 0)
    return valores.get('Caminhonete', 0)

def configurar_driver(headless=True):
    chrome_options = Options()
    if headless: 
        chrome_options.add_argument("--headless")
    
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    
    app_state = {"recentDestinations": [{"id": "Save as PDF", "origin": "local", "account": ""}], "selectedDestinationId": "Save as PDF", "version": 2}
    chrome_options.add_experimental_option("prefs", {"printing.print_preview_sticky_settings.appState": json.dumps(app_state), "savefile.default_directory": PASTA_DOWNLOADS})
    try: return webdriver.Chrome(options=chrome_options)
    except Exception as e:
        logging.critical(f"Falha Selenium: {e}")
        return None

def extrair_km_do_mapa(driver):
    try:
        xpath_prioritario = SELECTORS["google_maps"]["km_xpaths_list"][0]
        element = WebDriverWait(driver, 8).until(ec.presence_of_element_located((By.XPATH, xpath_prioritario)))
        texto = element.text.strip()
        if any(char.isdigit() for char in texto) and "km" in texto.lower():
            match = re.search(r"([\d\.]+)(?:,(\d+))?", texto)
            if match:
                base = match.group(1).replace('.', '')
                decimal = match.group(2) if match.group(2) else "0"
                km_final = float(f"{base}.{decimal}")
                logging.info(f"KM encontrado: {km_final}")
                return km_final, str(int(km_final))
    except:
        try:
            fallback = driver.find_element(By.XPATH, "//div[contains(text(), 'km')]")
            if "min" not in fallback.text:
                texto = fallback.text.strip()
                match = re.search(r"([\d\.]+)(?:,(\d+))?", texto)
                if match:
                    base = match.group(1).replace('.', '')
                    decimal = match.group(2) if match.group(2) else "0"
                    km_final = float(f"{base}.{decimal}")
                    logging.info(f"KM Fallback: {km_final}")
                    return km_final, str(int(km_final))
        except: pass
    return None, "KM_NAO_ENCONTRADO"

def get_valor_por_range(categoria, km_numerico):
    if km_numerico is None: return "VALOR_PENDENTE"
    categoria_limpa = str(categoria).strip().lower()
    ranges_da_categoria = VALOR_RANGES.get(categoria_limpa, [])
    for limite_km, valor in ranges_da_categoria:
        if km_numerico <= limite_km: return valor
    return "VALOR_NAO_ENCONTRADO"

def gerar_pdf_mapa(driver, nome_arquivo_pdf):
    try:
        WebDriverWait(driver, 20).until(ec.visibility_of_element_located((By.TAG_NAME, "canvas")))
        time.sleep(1)
        result = driver.execute_cdp_cmd("Page.printToPDF", {"landscape": False, "displayHeaderFooter": True, "printBackground": True, "marginTop": 1, "marginBottom": 1, "marginLeft": 0.5, "marginRight": 0.5})
        caminho_completo = os.path.join(PASTA_DOWNLOADS, nome_arquivo_pdf)
        with open(caminho_completo, "wb") as f: f.write(base64.b64decode(result['data']))
        return caminho_completo
    except: return None

def processar_mapa_single_instance(driver, placa, contrato, categoria, url, tipo, data):
    try:
        driver.get(url)
        km, km_str = extrair_km_do_mapa(driver)
        val = get_valor_por_range(categoria, km)
        if val == "VALOR_NAO_ENCONTRADO": return False, None, None, f"KM {km} fora range"
        nome = f"{placa}_{contrato}_{data}_{km_str}_{val}_{'REMO' if tipo == 'Remocao' else 'REST'}.pdf"
        pdf = gerar_pdf_mapa(driver, nome)
        if not pdf: return False, None, None, "Falha PDF"
        return True, km_str, val, pdf
    except Exception as e: return False, None, None, str(e)

def fazer_login_banco(driver):
    try:
        logging.info("Tentando abrir SafeDoc...")
        driver.get(URL_BANCO)
        driver.maximize_window()
        WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.XPATH, SELECTORS["login"]["usuario"])))
        driver.find_element(By.XPATH, SELECTORS["login"]["usuario"]).send_keys(USUARIO_BANCO)
        driver.find_element(By.XPATH, SELECTORS["login"]["senha"]).send_keys(SENHA_BANCO)
        driver.find_element(By.XPATH, SELECTORS["login"]["botao"]).click()
        
        logging.info("Aguardando menu principal...")
        WebDriverWait(driver, 30).until(ec.element_to_be_clickable((By.XPATH, SELECTORS["gca_menu"]["link_1"])))
        
        time.sleep(5) 
        
        return True
    except Exception as e:
        logging.error(f"Erro detalhado login: {e}")
        return False

def navegar_menu_gca(driver):
    try:
        logging.info("Navegando no Menu...")
        wait = WebDriverWait(driver, 30)
        
        el1 = wait.until(ec.element_to_be_clickable((By.XPATH, SELECTORS["gca_menu"]["link_1"])))
        el1.click()
        
        el2 = wait.until(ec.element_to_be_clickable((By.XPATH, SELECTORS["gca_menu"]["link_2"])))
        el2.click()
        
        el3 = wait.until(ec.element_to_be_clickable((By.XPATH, SELECTORS["gca_menu"]["link_3"])))
        el3.click()
        return True
    except Exception as e:
        logging.error(f"Erro ao navegar no menu: {e}")
        return False

def preencher_formulario_com_upload(driver, dados_upload, texto_ant=None):
    try:
        WebDriverWait(driver, 20).until(ec.element_to_be_clickable((By.XPATH, SELECTORS["form_upload"]["select_status"])))
        upload_element = driver.find_element(By.ID, SELECTORS["form_upload"]["input_arquivo"])
        driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.visibility = 'visible'; arguments[0].style.opacity = 1; arguments[0].style.height = '1px'; arguments[0].style.width = '1px';", upload_element)
        time.sleep(1)
        upload_element.send_keys(dados_upload['caminho_pdf'])
        driver.execute_script("validate(arguments[0]);", upload_element)
        time.sleep(1)
        
        Select(driver.find_element(By.XPATH, SELECTORS["form_upload"]["select_status"])).select_by_visible_text("Cadastrar")
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_data"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_data"]).send_keys(dados_upload['data'])
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_contrato"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_contrato"]).send_keys(dados_upload['contrato'])
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_placa"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_placa"]).send_keys(dados_upload['placa'].replace('-', ''))
        Select(driver.find_element(By.XPATH, SELECTORS["form_upload"]["select_tipo_despesa"])).select_by_visible_text("018 - GUINCHO")
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_valor"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_valor"]).send_keys(f"{dados_upload['valor']},00")
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_caixa_arquivo"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_caixa_arquivo"]).send_keys("0")
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_observacao"]).clear()
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["input_observacao"]).send_keys(dados_upload['tipo_str'])
        
        time.sleep(1)
        driver.find_element(By.XPATH, SELECTORS["form_upload"]["botao_salvar"]).click()
        
        def check_msg(d):
            try:
                txt = d.find_element(By.XPATH, SELECTORS["form_upload"]["mensagem_sucesso"]).text.strip()
                return txt if txt and txt != texto_ant else False
            except: return False
            
        res = WebDriverWait(driver, 20).until(check_msg)
        return True, res
    except Exception as e:
        return False, str(e)

def enviar_resumo_telegram(sucesso, falha):
    token, chat = os.getenv("TELEGRAM_BOT_TOKEN"), os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat: return
    try:
        msg = ["--- 🤖 Resumo Automação ---"]
        if sucesso:
            msg.append("\n✅ SUCESSOS:")
            total = 0
            for item in sucesso:
                msg.append(f"\nPlaca: {item[COLUNA_PLACA]}")
                for k in ['valor_rem', 'valor_rest']:
                    try:
                        val = int(item.get(k, 0))
                        if val > 0: 
                            total += val
                            tipo = "Remoção" if k == 'valor_rem' else "Restituição"
                            msg.append(f"  • {tipo}: R$ {val},00")
                    except: pass
                if 'JPR' in str(item.get('Transportadora', '')).upper():
                    msg.append(f"  • (JPR): R$ {item.get('Valor_Base_Guincho2')}")
            msg.append(f"\n💰 Total: R$ {total},00")
        if falha:
            msg.append("\n\n❌ FALHAS:")
            for item in falha: msg.append(f"  • {item.get('placa', '?')}: {item.get('motivo', 'Erro')}")
            
        async def send(tk, cid, txt):
            await telegram.Bot(tk).send_message(chat_id=cid, text=txt)
        asyncio.run(send(token, chat, "\n".join(msg)))
    except: pass

def enviar_email_outlook(lista_uploads_sucesso):
    if not lista_uploads_sucesso:
        logging.info("Nenhum upload realizado, e-mail não será enviado.")
        return

    destinatario = os.getenv("EMAIL_FINANCEIRO")
    
    if not destinatario:
        logging.warning("ABORTADO: Variável 'EMAIL_FINANCEIRO' não encontrada no .env.")
        return

    try:
        try: diretorio_script = os.path.dirname(os.path.abspath(__file__))
        except: diretorio_script = os.getcwd()
        
        hoje_str_log = datetime.date.today().strftime("%Y-%m-%d")
        pasta_logs_hoje = os.path.join(diretorio_script, ".logs", hoje_str_log)
        
        qtd_logs = len(glob.glob(os.path.join(pasta_logs_hoje, "log_*.txt")))
        versao = qtd_logs if qtd_logs > 0 else 1
    except:
        versao = 1

    data_hoje_formatada = datetime.date.today().strftime('%d/%m/%Y')
    assunto = f"Relatório de Reembolsos - Processamento v{versao} {data_hoje_formatada}"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #eeeeee; margin: 0; padding: 20px; }}
        .container {{ max-width: 700px; margin: 0 auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.15); border-top: 5px solid #CC0000; }}
        .header {{ background-color: #ffffff; color: #333; padding: 20px; text-align: center; border-bottom: 1px solid #eee; }}
        .header h2 {{ margin: 0; font-size: 22px; font-weight: 700; color: #CC0000; text-transform: uppercase; letter-spacing: 0.5px; }}
        .content {{ padding: 30px; color: #444; line-height: 1.6; }}
        .table-container {{ margin-top: 25px; margin-bottom: 25px; border: 1px solid #e0e0e0; border-radius: 6px; overflow: hidden; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
        th {{ background-color: #f2f2f2; color: #222; font-weight: bold; text-align: left; padding: 12px; border-bottom: 2px solid #ccc; }}
        td {{ padding: 12px; border-bottom: 1px solid #f0f0f0; color: #555; vertical-align: middle; }}
        
        /* Tags de Status */
        .tag-remo {{ background-color: #e0e0e0; color: #333; padding: 5px 10px; border-radius: 4px; font-size: 11px; font-weight: bold; text-transform: uppercase; }}
        .tag-rest {{ background-color: #ffebee; color: #c62828; padding: 5px 10px; border-radius: 4px; font-size: 11px; font-weight: bold; text-transform: uppercase; border: 1px solid #ffcdd2; }}
        
        .total-row {{ background-color: #222222; color: #ffffff; font-weight: bold; }}
        .footer {{ background-color: #f9f9f9; padding: 20px; text-align: center; font-size: 11px; color: #999; border-top: 1px solid #eee; }}
        .signature {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; }}
        .dev-name {{ font-weight: bold; color: #CC0000; font-size: 16px; }}
        .dev-role {{ color: #666; font-size: 14px; margin-top: 2px; display: block; }}
        .bot-badge {{ display: inline-block; background-color: #000; color: #fff; padding: 2px 6px; border-radius: 3px; font-size: 10px; margin-right: 5px; }}
    </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>Relatório de Reembolsos</h2>
            </div>
            
            <div class="content">
                <p>Olá, <strong>Equipe Financeira</strong>.</p>
                <p>O <strong>Bot Restituição</strong> finalizou a rodada de processamento <strong>v{versao}</strong> referente a data de hoje ({data_hoje_formatada}).</p>
                <p>Abaixo estão consolidados os valores lançados com sucesso no sistema SafeDOC:</p>

                <div class="table-container">
                    <table>
                        <thead>
                            <tr>
                                <th>Placa</th>
                                <th>Contrato</th>
                                <th>Data Fato</th>
                                <th>Tipo</th>
                                <th>Valor (R$)</th>
                            </tr>
                        </thead>
                        <tbody>
    """
    
    total_valor = 0
    for item in lista_uploads_sucesso:
        try:
            val_float = float(item['valor'])
            val_fmt = f"{val_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            total_valor += val_float
        except:
            val_fmt = item['valor']
        
        tipo_clean = "Restituição" if "Restituicao" in item['tipo_str'] else "Remoção"
        classe_tag = "tag-rest" if "Restituicao" in item['tipo_str'] else "tag-remo"
            
        html_body += f"""
            <tr>
                <td style="font-family: monospace; font-size: 13px; font-weight: 600;">{item['placa']}</td>
                <td>{item['contrato']}</td>
                <td>{item['data']}</td>
                <td><span class="{classe_tag}">{tipo_clean}</span></td>
                <td>R$ {val_fmt}</td>
            </tr>
        """

    total_fmt = f"{total_valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    html_body += f"""
                        <tr class="total-row">
                            <td colspan="4" style="text-align: right; padding-right: 20px;">TOTAL LANÇADO:</td>
                            <td>R$ {total_fmt}</td>
                        </tr>
                        </tbody>
                    </table>
                </div>

                <div class="signature">
                    <p>Atenciosamente,</p>
                    <span class="dev-name">Vinícius Lima</span>
                    <span class="dev-role">Loop Transportes</span>
                </div>
            </div>

            <div class="footer">
                <span class="bot-badge">BOT</span> Relatório gerado automaticamente pelo <strong>Bot Restituição, Loop Transportes</strong><br>
                Processamento realizado em {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')}
            </div>
        </div>
    </body>
    </html>
    """

    try:
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = destinatario
        mail.Subject = assunto
        mail.HTMLBody = html_body
        mail.Send()
        logging.info(f"E-mail enviado para {destinatario} | Assunto: {assunto}")
    except Exception as e:
        logging.error(f"Erro ao enviar e-mail via Outlook: {e}")

def aplicar_calculos_analise(df, lista_placas_processadas=None):
    try:
        logging.info("Aplicando cálculos de análise...")
        
        cols_num = ['Valor_Base_Guincho', 'valor_rem', 'Valor_Base_Guincho2']
        for col in cols_num:
            if col not in df.columns:
                df[col] = 0.0
            else:
                df[col] = df[col].astype(str).str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        if 'Calculo_cobrança' not in df.columns:
            df['Calculo_cobrança'] = 0.0

        if lista_placas_processadas:
            filtro = df[COLUNA_PLACA].isin(lista_placas_processadas)
        else:
            filtro = slice(None)

        if 'Tipo_Restituicao' in df.columns:
            df.loc[filtro, 'Teste'] = df.loc[filtro, 'Tipo_Restituicao'].astype(str).str.strip().apply(
                lambda x: 1 if x == "Transportadora" else 0
            )
        
        def calc_cobranca(row):
            try:
                tipo_lib = str(row.get('Tipo_Liberacao', '')).strip()
                teste = row.get('Teste', 0)
                v_base = float(row.get('Valor_Base_Guincho', 0))
                v_rem = float(row.get('valor_rem', 0))
                v_base2 = float(row.get('Valor_Base_Guincho2', 0))

                resultado = 0.0
                if tipo_lib == "Acordo" and teste == 1:
                    resultado = ((v_base - v_rem) + v_base2) * 1.15
                elif tipo_lib == "Acordo" and teste == 0:
                    resultado = (v_base - v_rem) * 1.15
                
                return max(0.0, resultado)
            except:
                return 0.0

        df.loc[filtro, 'Calculo_cobrança'] = df.loc[filtro].apply(calc_cobranca, axis=1)
        
        return df

    except Exception as e:
        logging.error(f"Erro ao aplicar cálculos de análise: {e}")
        return df

def salvar_historico_parcial(res_final):
    try:
        if not res_final: return
        
        try:
            df_hist = pd.read_excel(NOME_ARQUIVO_HISTORICO)
            df_hist[COLUNA_PLACA] = df_hist[COLUNA_PLACA].astype(str).str.strip()
        except:
            df_hist = pd.DataFrame(list(res_final.values()))
            df_hist = aplicar_calculos_analise(df_hist, lista_placas_processadas=None)
            
        cols_to_fix = [
            'km_remocao', 'valor_rem', 'km_restituicao', 'valor_rest', 
            'Valor_Base_Guincho2', 'Valor_Base_Guincho',
            'Tipo_Liberacao', 'Tipo_Restituicao', 
            'Transportadora', 'Contrato_Externo',
            'Conferencia SafeDoc'
        ]
        
        for col in cols_to_fix:
            if col not in df_hist.columns: df_hist[col] = None
            df_hist[col] = df_hist[col].astype('object')

        placas_existentes = df_hist[COLUNA_PLACA].unique()
        novas_linhas = [dados for p, dados in res_final.items() if p not in placas_existentes]
        if novas_linhas:
            df_hist = pd.concat([df_hist, pd.DataFrame(novas_linhas)], ignore_index=True)
            for col in cols_to_fix: 
                if col in df_hist.columns: df_hist[col] = df_hist[col].astype('object')

        df_hist.set_index(COLUNA_PLACA, inplace=True)
        updates_count = 0
        
        for placa, dados in res_final.items():
            if placa in df_hist.index:
                campos = [
                    'km_remocao', 'valor_rem', 'km_restituicao', 'valor_rest', 
                    'Valor_Base_Guincho2', 'Valor_Base_Guincho', 'Tipo_Liberacao', 
                    'Tipo_Restituicao', 'Conferencia SafeDoc'
                ]
                for campo in campos:
                    val = dados.get(campo)
                    if val is not None: df_hist.at[placa, campo] = val
                
                updates_count += 1
        
        df_hist.reset_index(inplace=True)
        
        placas_da_rodada = [str(k).strip() for k in res_final.keys()]
        df_hist = aplicar_calculos_analise(df_hist, lista_placas_processadas=placas_da_rodada)

        cols_data = ['Data de Remoção', 'Data Restituição', 'Fechamento Solicitação']
        for col in cols_data:
            if col in df_hist.columns: df_hist[col] = df_hist[col].apply(formatar_data_ptbr)

        cols_moeda = ['valor_rem', 'valor_rest', 'Valor_Base_Guincho', 'Valor_Base_Guincho2', 'Calculo_cobrança']
        for col in cols_moeda:
            if col in df_hist.columns:
                df_hist[col] = df_hist[col].apply(formatar_moeda_br)

        df_hist.to_excel(NOME_ARQUIVO_HISTORICO, index=False)
        logging.info(f"CHECKPOINT: Histórico atualizado ({updates_count} placas).")
        
    except Exception as e:
        logging.error(f"Erro ao salvar checkpoint: {e}")

def atualizar_planilha_base_status(lista_placas_sucesso):
    if not lista_placas_sucesso: return
    
    logging.info("💾 Atualizando Status na planilha Base (Preservando formatação)...")
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(NOME_ARQUIVO_EXCEL)
        
        if NOME_ABA_CALCULOS in wb.sheetnames:
            ws = wb[NOME_ABA_CALCULOS]
        else:
            logging.error(f"❌ Aba '{NOME_ABA_CALCULOS}' não encontrada na Base.")
            return

        col_placa_idx = None
        col_status_idx = None
        header_row_idx = 1

        found_header = False
        for r in range(1, 6): 
            row_values = [c.value for c in ws[r]]
            row_strs = [str(v).strip() if v else "" for v in row_values]
            
            if COLUNA_PLACA in row_strs and COLUNA_STATUS_SAFEDOC in row_strs:
                header_row_idx = r
                for idx, val in enumerate(row_strs):
                    if val == COLUNA_PLACA: col_placa_idx = idx + 1
                    elif val == COLUNA_STATUS_SAFEDOC: col_status_idx = idx + 1
                found_header = True
                break
        
        if not found_header or not col_placa_idx or not col_status_idx:
            logging.error("❌ Colunas 'Placa' ou 'Conferencia SafeDoc' não encontradas na Base.")
            return

        count = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            cell_placa = row[col_placa_idx - 1] 
            cell_status = row[col_status_idx - 1]
            
            placa_val = str(cell_placa.value).strip()
            
            if placa_val in lista_placas_sucesso:
                status_atual = str(cell_status.value).strip().upper()
                if status_atual not in ["FATURADO", "APROVADO"]:
                    cell_status.value = "Aprovado"
                    count += 1
        
        if count > 0:
            wb.save(NOME_ARQUIVO_EXCEL)
            logging.info(f"✅ Base atualizada: {count} placas marcadas como 'Aprovado'.")
        else:
            logging.info("ℹ️ Nenhuma alteração necessária na Base.")
            
    except Exception as e:
        logging.error(f"❌ Erro ao atualizar Excel Base (openpyxl): {e}")

# --- MAIN ---
def iniciar_automacao_completa():
    configurar_logger_dinamico()
    logging.info("--- Automação Restituição ---")

    dict_transp, dict_patio = carregar_bases_de_enderecos()
    df_ext = carregar_base_externa_rede()
    tabela_jpr = carregar_tabela_custos_jpr()

    if os.path.exists(NOME_ARQUIVO_HISTORICO):
        try:
            df_hist = pd.read_excel(NOME_ARQUIVO_HISTORICO)
        except:
            df_hist = pd.DataFrame(columns=[COLUNA_PLACA])
    else:
        df_hist = pd.DataFrame(columns=[COLUNA_PLACA])
    
    df_hist = sincronizar_dados_dinamicos_local(df_hist, df_ext)
    
    cols_data = ['Data de Remoção', 'Data Restituição', 'Fechamento Solicitação']
    for col in cols_data:
        if col in df_hist.columns:
            df_hist[col] = df_hist[col].apply(formatar_data_ptbr)
            
    df_hist.to_excel(NOME_ARQUIVO_HISTORICO, index=False)
    logging.info("Histórico Sincronizado e Salvo.")

    try:
        df = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=NOME_ABA_CALCULOS)
        if COLUNA_PLACA in df.columns: df[COLUNA_PLACA] = df[COLUNA_PLACA].astype(str).str.strip()
        
        if not df_hist.empty and COLUNA_TESTE in df_hist.columns:
            df_teste_ref = df_hist[[COLUNA_PLACA, COLUNA_TESTE]].copy()
            df_teste_ref[COLUNA_PLACA] = df_teste_ref[COLUNA_PLACA].astype(str).str.strip()
            df_teste_ref.drop_duplicates(subset=[COLUNA_PLACA], inplace=True)
            
            if COLUNA_TESTE in df.columns:
                df.drop(columns=[COLUNA_TESTE], inplace=True)

            df = pd.merge(df, df_teste_ref, on=COLUNA_PLACA, how='left')
            df[COLUNA_TESTE] = pd.to_numeric(df[COLUNA_TESTE], errors='coerce').fillna(0).astype(int)
        else:
            df[COLUNA_TESTE] = 0

        if not df_ext.empty:
            df = pd.merge(df, df_ext, left_on=COLUNA_PLACA, right_on='placa_key', how='left')

        for i, row in df.iterrows():
            tn = str(row.get('transp_raw', '')).strip().upper()
            df.at[i, COLUNA_END1] = dict_transp.get(tn, "")
            pn = limpar_texto_estilo_excel(str(row.get('patio_raw', '')))
            df.at[i, COLUNA_END2] = dict_patio.get(pn, "")
            cn = str(row.get('cidade_raw', ''))
            df.at[i, COLUNA_END3] = limpar_texto_estilo_excel(cn) if cn != 'nan' else ""
            
    except Exception as e:
        logging.critical(f"Erro ao preparar dados: {e}")
        return

    try:
        hist_check = pd.read_excel(NOME_ARQUIVO_HISTORICO)
        for c in ['valor_rem', 'km_remocao']:
            if c in hist_check.columns: hist_check[c] = pd.to_numeric(hist_check[c], errors='coerce').fillna(0)
            else: hist_check[c] = 0
        
        concluidas = hist_check[ (hist_check['valor_rem'] > 0) | (hist_check['km_remocao'] > 0) ][COLUNA_PLACA].astype(str).str.strip().tolist()
    except: 
        concluidas = []

    logging.info("Iniciando processamento...")
    
    res_final = {}
    uploads = []
    
    driver = configurar_driver(headless=True)
    if not driver:
        return

    dt_hoje = datetime.date.today().strftime("%d-%m-%Y")
    count_processados = 0

    for idx, row in df.iterrows():
        placa = str(row[COLUNA_PLACA]).strip()
        if not placa or placa == 'nan': continue

        status_safedoc = str(row.get(COLUNA_STATUS_SAFEDOC, '')).strip().upper()
        if status_safedoc in ['NAN', 'NONE']: status_safedoc = ""

        executar_remo = False
        executar_rest = False
        motivo_acao = ""
        
        if status_safedoc in ["APROVADO", "FATURADO"]:
            continue 
        
        elif status_safedoc in ["NEGADO", "DEVOLVIDO"]:
            executar_remo = True
            executar_rest = True
            motivo_acao = f"Forçado por Status {status_safedoc}"
            
        elif status_safedoc == "PENDENTE REMO":
            executar_remo = True
            executar_rest = False
            motivo_acao = "Forçado Pendente Remo"
            
        elif status_safedoc == "PENDENTE REST":
            executar_remo = False
            executar_rest = True
            motivo_acao = "Forçado Pendente Rest"
            
        else:
            if placa not in concluidas:
                executar_remo = True
                executar_rest = True 
                motivo_acao = "Novo Processo (Status Vazio)"
            else:
                continue

        flag_teste = row.get(COLUNA_TESTE, 0)
        
        if executar_rest and flag_teste == 0:
            executar_rest = False 
            if "PENDENTE REST" in status_safedoc:
                logging.warning(f"Placa {placa}: Status pede Restituição, mas flag Teste=0 impede.")

        if not executar_remo and not executar_rest:
            continue

        count_processados += 1
        logging.info(f"Processando {placa} | Ação: {motivo_acao} | Remo: {executar_remo}, Rest: {executar_rest}")

        contrato_safe = row.get('contrato_externo') or row.get('Contrato') or "S_CONTRATO"
        categoria_safe = row.get('Categoria_Ext') or row.get('Categoria') or "Leve"
        transp_atual = str(row.get('transp_raw', '')).strip()
        val_orig_db = row.get('valor_base_db', 0)
        
        end1, end2, end3 = row[COLUNA_END1], row[COLUNA_END2], row[COLUNA_END3]

        if placa not in res_final:
            res_final[placa] = {
                COLUNA_PLACA: placa, 'valor_rem': 0, 'km_remocao': 0, 
                'valor_rest': 0, 'km_restituicao': 0, 'falhas': [],
                'Contrato_Externo': contrato_safe, 'Valor_Base_Guincho': val_orig_db,
                'Transportadora': transp_atual, 'Valor_Base_Guincho2': 0,
                'Tipo_Liberacao': row.get('Tipo de liberação'),  
                'Tipo_Restituicao': row.get('Tipo de restituição')
            }

        if not end1 or len(str(end1)) < 3:
            res_final[placa]['falhas'].append("Endereço Transp Inválido")
            continue

        url_rem = f"https://www.google.com/maps/dir/{str(end1).replace(' ','+')}/{str(end2).replace(' ','+')}/{str(end3).replace(' ','+')}/{str(end1).replace(' ','+')}"
        url_rest = f"https://www.google.com/maps/dir/{str(end1).replace(' ','+')}/{str(end3).replace(' ','+')}/{str(end1).replace(' ','+')}/{str(end2).replace(' ','+')}"

        remo_ok = False
        val_remo_final = 0.0

        if executar_remo:
            ok, km, val, pdf = processar_mapa_single_instance(driver, placa, contrato_safe, categoria_safe, url_rem, "Remocao", dt_hoje)
            if ok:
                res_final[placa]['valor_rem'] = val
                res_final[placa]['km_remocao'] = km
                uploads.append({'placa': placa, 'contrato': contrato_safe, 'data': dt_hoje, 'valor': str(val), 'tipo_str': "Remocao", 'caminho_pdf': pdf})
                val_remo_final = val
                remo_ok = True
            else:
                logging.error(f"Falha Maps Remo ({placa}): {pdf}")
                res_final[placa]['falhas'].append(f"Maps Remo: {pdf}")
        else:
            remo_ok = True 
            try: val_remo_final = float(res_final[placa].get('valor_rem', 0))
            except: val_remo_final = 0.0

        res_final[placa]['Valor_Base_Guincho2'] = calcular_valor_restituicao_final(transp_atual, row.get('cidade_raw', ''), row.get('patio_raw', ''), categoria_safe, val_orig_db, tabela_jpr)
        v_cobranca = calcular_cobranca_individual(row.get('Tipo_Liberacao'), row.get('Tipo_Restituicao'), val_orig_db, val_remo_final, res_final[placa]['Valor_Base_Guincho2'])
        res_final[placa]['Calculo_cobrança'] = v_cobranca

        if executar_rest:
            if remo_ok: 
                if not end3:
                    res_final[placa]['falhas'].append("Sem Cidade Destino")
                else:
                    ok2, km2, val2, pdf2 = processar_mapa_single_instance(driver, placa, contrato_safe, categoria_safe, url_rest, "Restituicao", dt_hoje)
                    if ok2:
                        res_final[placa]['valor_rest'] = val2
                        res_final[placa]['km_restituicao'] = km2
                        uploads.append({'placa': placa, 'contrato': contrato_safe, 'data': dt_hoje, 'valor': str(val2), 'tipo_str': "Restituicao", 'caminho_pdf': pdf2})
                    else:
                        res_final[placa]['falhas'].append(f"Maps Rest: {pdf2}")
            else:
                logging.warning(f"Restituição pulada para {placa} pois Remoção falhou.")

    driver.quit()
    
    if count_processados == 0:
        logging.info("Nenhuma placa elegível para processamento nesta rodada.")
        return

    salvar_historico_parcial(res_final)

    logging.info(f"Fase Banco: {len(uploads)} uploads pendentes.")
    uploads_confirmados = [] 

    if uploads:
        driver_banco = configurar_driver(headless=True)
        if driver_banco and fazer_login_banco(driver_banco):
            try:
                if navegar_menu_gca(driver_banco):
                    WebDriverWait(driver_banco, 10).until(ec.frame_to_be_available_and_switch_to_it((By.ID, SELECTORS["iframes"]["externo"])))
                    WebDriverWait(driver_banco, 10).until(ec.frame_to_be_available_and_switch_to_it((By.ID, SELECTORS["iframes"]["interno"])))
                    
                    last_txt = None
                    for d in uploads:
                        p = d['placa']
                        logging.info(f"Iniciando Upload: {p} ({d['tipo_str']})")
                        ok, txt = preencher_formulario_com_upload(driver_banco, d, last_txt)
                        if ok:
                            logging.info(f"Upload OK: {p}")
                            uploads_confirmados.append(d) 
                            last_txt = txt
                        else:
                            logging.error(f"Falha Upload {p}: {txt}")
                            res_final[p]['falhas'].append(f"Banco {d['tipo_str']}: {txt}")
                else:
                    logging.error("Falha ao navegar no Menu.")
                    for d in uploads: res_final[d['placa']]['falhas'].append("Erro Menu Banco")
            finally:
                driver_banco.quit()
        else:
            logging.critical("Falha ao abrir SafeDoc ou fazer login.")
            for d in uploads: res_final[d['placa']]['falhas'].append("Erro Geral Login Banco")

    if res_final:
        try: df_atual = pd.DataFrame(list(res_final.values()))
        except: pass

    salvar_historico_parcial(res_final)

    sucessos = [d for d in res_final.values() if not d['falhas']]
    falhas = [{'placa': k, 'motivo': v['falhas']} for k, v in res_final.items() if v['falhas']]
    
    enviar_resumo_telegram(sucessos, falhas)
    #enviar_email_outlook(uploads_confirmados)

    placas_para_aprovar = [item['placa'] for item in uploads_confirmados]
    placas_para_aprovar = list(set(placas_para_aprovar))
    
    if placas_para_aprovar:
        logging.info("Atualizando status 'Aprovado' no Histórico e na Base...")
        
        for p in placas_para_aprovar:
            if p in res_final:
                res_final[p]['Conferencia SafeDoc'] = "Aprovado"
        
        salvar_historico_parcial(res_final)

        atualizar_planilha_base_status(placas_para_aprovar)

    logging.info("FIM DO PROCESSO.")

if __name__ == "__main__":
    iniciar_automacao_completa()